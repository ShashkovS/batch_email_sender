# -*- coding: utf-8 -*-
import importlib.util

modules_to_check = [
    'PyQt5',
    'openpyxl',
    'keyring'
]

if not all(importlib.util.find_spec(name) for name in modules_to_check):
    import ensurepip
    ensurepip.bootstrap(upgrade=False, user=True)
    import pip
    for name in modules_to_check:
        pip.main(['install', "--user", name])


# All imports
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate, formataddr
from os.path import basename
from typing import List
from email.mime.multipart import MIMEMultipart
import smtplib
import keyring
import openpyxl
import os
import subprocess
import traceback
import re
from email.mime.application import MIMEApplication
from email.header import Header
import sys
import queue
from PyQt5 import QtCore, QtWidgets
from PyQt5.Qt import *



###### files_parsers ######
ONE_PLUS_ONE_FOR_HEADER = 2
OK_COLUMN = 1
OKOK = 'ok'
ORIGINAL_ROW_NUM = 'row_num_WcCRve89'
EMAIL_REGEX = r"\s*([a-zA-Z0-9'_][a-zA-Z0-9'._+-]{,63}@[a-zA-Z0-9.-]{,254}[a-zA-Z0-9])\s*"

def rtv_table(xls_name):
    try:
        xl_workbook = openpyxl.load_workbook(filename=xls_name, read_only=True, data_only=True)
    except FileNotFoundError:
        raise Exception(f'Файл {xls_name} не найден')
    xl_sheet_names = xl_workbook.sheetnames
    xl_sheet = xl_workbook[xl_sheet_names[0]]
    columns = []
    bold_columns = []
    row_iter = iter(xl_sheet.rows)
    for cell in next(row_iter):
        title = str(cell.value)
        if title:
            columns.append(title) # cell.column
            if cell.font.bold:
                bold_columns.append(title)
    row_dict = {title: '' for title in columns}
    rows_list = []
    for rn, row in enumerate(row_iter, start=2):
        cur = row_dict.copy()
        cur[ORIGINAL_ROW_NUM] = rn
        for col_name, cell in zip(columns, row):
            cur[col_name] = str(cell.value).replace('None', '')
        cur['email'] = re.findall(EMAIL_REGEX, cur['email'])
        rows_list.append(cur)
    return rows_list, bold_columns


def set_ok(xls_name, row_num_real):
    while True:
        try:
            xl_workbook = openpyxl.load_workbook(filename=xls_name, read_only=False, data_only=False)
            break
        except FileNotFoundError:
            raise Exception(f'Файл {xls_name} не найден')
        except PermissionError:
            raise Exception(f'Файл {xls_name} заблокирован. Сохраните и закойте его. В него будут вноситься отметки об успешности отправки')
            # continue
    xl_sheet_names = xl_workbook.get_sheet_names()
    xl_sheet = xl_workbook.get_sheet_by_name(xl_sheet_names[0])
    xl_sheet.cell(row=row_num_real, column=OK_COLUMN).value = OKOK
    xl_workbook.save(filename=xls_name)


def rtv_template(template_name):
    try:
        with open(template_name, encoding='utf-8') as f:
            template = f.read()
    except FileNotFoundError:
        raise Exception(f'Файл с шаблоном {template_name} не найден')
    return template


def rtv_table_and_template(xls_name, template_name):
    rows_list, bold_columns = rtv_table(xls_name)
    template = rtv_template(template_name)
    if not rows_list:
        raise Exception(f'В файле {xls_name} не обнаружены строки с данными')
    first_data_row = rows_list[0]
    # Проверяем обязательные столбцы: ok, email, subject
    for col_name in ['ok', 'email', 'subject']:
        if col_name not in first_data_row:
            raise Exception(f'В таблице {xls_name} обязательно должен быть столбец {col_name}')
    # Проверяем, что есть всё, что указано в шаблоне
    try:
        template.format(**first_data_row)
    except KeyError as e:
        raise Exception(f'В таблице {xls_name} должен быть столбец {e!s}, так как он упоминается в шаблоне {template_name}')
    # Теперь проверяем существование всех вложений
    attach_cols = [key for key in first_data_row if key.startswith('attach')]
    for rn, row in enumerate(rows_list):
        if attach_cols:
            if not row['email']:
                continue
            for attach_key in attach_cols:
                attach_name = row[attach_key]
                if attach_name and not os.path.isfile(attach_name):
                    raise Exception(f'В таблице {xls_name} в строчке {rn+ONE_PLUS_ONE_FOR_HEADER} в столбце {attach_key} указано вложение "{attach_name}". Этот файл не найден')
            row['attach_list'] = [row[attach_key] for attach_key in attach_cols]
        else:
            row['attach_list'] = []
    # Проверили, что всё работает. Проверили, что вложения существуют
    return rows_list, bold_columns, template




###### email_stuff ######

class EmailEnvelope:
    def __init__(self, smtp_server, login, password, sender_addr, sender_name='', copy_addrs=None):
        """Сохраняем всем параметры в атрибутах. Больше ничего не делаем"""
        self.smtp_server = smtp_server
        self.login = login
        self.password = password
        self.sender_addr = sender_addr
        self.sender_name = sender_name
        self.copy_addrs = copy_addrs or []
        self.smtp = None
        self.send_queue = queue.Queue()

    def connect_to_server(self):
        """Подключаемся к серверу"""
        if self.smtp is None:
            self.smtp = smtplib.SMTP()
            # self.smtp.set_debuglevel(1)
        # Проверяем подключение
        try:
            status = self.smtp.noop()[0]
        except smtplib.SMTPServerDisconnected as e:
            status = -1
        # Если не ОК, то переподключаемся
        if status != 250:
            self.smtp.connect(host=self.smtp_server)
            self.smtp.ehlo_or_helo_if_needed()
            self.smtp.starttls()
            self.smtp.login(self.login, self.password)

    def verify_credentials(self):
        self.connect_to_server()
        try:
            status, _ = self.smtp.noop()
        except smtplib.SMTPServerDisconnected as e:
            status = -1
        if status != 250:
            raise Exception(f'Не удалось подключиться для отправки почты с адреса {self.login}')

    def add_mail_to_queue(self, recipients: List[str], subject, html, files=None, xls_id=None, qt_id=None):
        msg = MIMEMultipart()
        msg['From'] = formataddr((Header(self.sender_name, 'utf-8').encode(), self.sender_addr))
        msg['To'] = COMMASPACE.join(recipients)
        msg['Cc'] = COMMASPACE.join(self.copy_addrs)
        msg['Date'] = formatdate(localtime=True)
        msg['Subject'] = Header(subject, 'utf-8')
        msg.attach(MIMEText(html.encode('utf-8'), 'html', 'utf-8'))
        for f in files or []:
            if not f:
                continue
            with open(f, "rb") as fil:
                part = MIMEApplication(
                    fil.read(),
                    Name=basename(f)
                )
            # After the file is closed
            part['Content-Disposition'] = 'attachment; filename="%s"' % basename(f)
            msg.attach(part)
        mail = dict(from_addr=self.sender_addr, to_addrs=recipients + self.copy_addrs, msg=msg.as_string(),
                    xls_id=xls_id, qt_id=qt_id)
        self.send_queue.put(mail)

    def send_next(self):
        try:
            mail = self.send_queue.get(block=False)
        except queue.Empty as e:
            raise StopIteration
        self.connect_to_server()
        self.smtp.sendmail(from_addr=mail['from_addr'], to_addrs=mail['to_addrs'], msg=mail['msg'])
        return mail

    def __copy__(self):
        return self.__class__(smtp_server=self.smtp_server, login=self.login, password=self.password,
                              sender_addr=self.sender_addr, sender_name=self.sender_name, copy_addrs=self.copy_addrs)

    def copy(self):
        return self.__copy__()




###### ui_email_and_passw ######

class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(400, 147)
        self.gridLayout = QtWidgets.QGridLayout(Dialog)
        self.gridLayout.setObjectName("gridLayout")
        self.formLayout = QtWidgets.QFormLayout()
        self.formLayout.setObjectName("formLayout")
        self.line_email = QtWidgets.QLineEdit(Dialog)
        self.line_email.setObjectName("line_email")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.line_email)
        self.label_1 = QtWidgets.QLabel(Dialog)
        self.label_1.setAlignment(QtCore.Qt.AlignLeading | QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self.label_1.setObjectName("label_1")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.label_1)
        self.label_2 = QtWidgets.QLabel(Dialog)
        self.label_2.setObjectName("label_2")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.label_2)
        self.line_password = QtWidgets.QLineEdit(Dialog)
        self.line_password.setEchoMode(QtWidgets.QLineEdit.Password)
        self.line_password.setObjectName("line_password")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.line_password)
        self.label_3 = QtWidgets.QLabel(Dialog)
        self.label_3.setObjectName("label_3")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.LabelRole, self.label_3)
        self.line_sender = QtWidgets.QLineEdit(Dialog)
        self.line_sender.setObjectName("line_sender")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.FieldRole, self.line_sender)
        self.label_4 = QtWidgets.QLabel(Dialog)
        self.label_4.setObjectName("label_4")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.LabelRole, self.label_4)
        self.line_smtpserver = QtWidgets.QLineEdit(Dialog)
        self.line_smtpserver.setObjectName("line_smtpserver")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.FieldRole, self.line_smtpserver)

        self.label_5 = QtWidgets.QLabel(Dialog)
        self.label_5.setObjectName("label_5")
        self.formLayout.setWidget(4, QtWidgets.QFormLayout.LabelRole, self.label_5)
        self.line_send_copy = QtWidgets.QLineEdit(Dialog)
        self.line_send_copy.setObjectName("line_send_copy")
        self.formLayout.setWidget(4, QtWidgets.QFormLayout.FieldRole, self.line_send_copy)


        self.label_6 = QtWidgets.QLabel(Dialog)
        self.label_6.setObjectName("label_6")
        self.formLayout.setWidget(5, QtWidgets.QFormLayout.LabelRole, self.label_6)

        self.save_passw_cb = QtWidgets.QCheckBox(Dialog)
        self.save_passw_cb.setObjectName("save_passw_cb")
        self.formLayout.setWidget(5, QtWidgets.QFormLayout.FieldRole, self.save_passw_cb)

        self.gridLayout.addLayout(self.formLayout, 0, 0, 1, 1)
        self.buttonBox = QtWidgets.QDialogButtonBox(Dialog)
        self.buttonBox.setOrientation(QtCore.Qt.Horizontal)
        self.buttonBox.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel|QtWidgets.QDialogButtonBox.Ok)
        self.buttonBox.setCenterButtons(True)
        self.buttonBox.setObjectName("buttonBox")
        self.gridLayout.addWidget(self.buttonBox, 1, 0, 1, 1)
        self.buttonBox.accepted.connect(Dialog.accept)
        self.buttonBox.rejected.connect(Dialog.reject)
        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Введите email, пароль и прочее"))
        self.label_1.setText(_translate("Dialog", "Логин (email)"))
        self.label_2.setText(_translate("Dialog", "Пароль"))
        self.label_3.setText(_translate("Dialog", "Отправитель"))
        self.label_4.setText(_translate("Dialog", "SMTP сервер"))
        self.label_5.setText(_translate("Dialog", "Поставить в копию"))
        self.label_6.setText(_translate("Dialog", "Сохранить пароль"))


###### ui_main_window ######


class Ui_MainWindow(object):
    def __init__(self):
        super().__init__()
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(691, 662)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.gridLayout_4 = QtWidgets.QGridLayout()
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.pushButton_ask_and_send = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_ask_and_send.setObjectName("pushButton_ask_and_send")
        self.gridLayout_4.addWidget(self.pushButton_ask_and_send, 0, 0, 1, 1)
        self.pushButton_cancel_send = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_cancel_send.setCheckable(False)
        self.pushButton_cancel_send.setObjectName("pushButton_cancel_send")
        self.gridLayout_4.addWidget(self.pushButton_cancel_send, 0, 1, 1, 1)
        self.listWidget_emails = QtWidgets.QListWidget(self.centralwidget)
        self.listWidget_emails.setObjectName("listWidget_emails")
        self.gridLayout_4.addWidget(self.listWidget_emails, 1, 0, 1, 2)
        self.gridLayout_2.addLayout(self.gridLayout_4, 0, 0, 1, 1)
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.pushButton_open_list_and_template = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_open_list_and_template.setObjectName("pushButton_open_list_and_template")
        self.gridLayout.addWidget(self.pushButton_open_list_and_template, 0, 0, 1, 1)
        self.textBrowser = QtWidgets.QTextBrowser(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(99)
        sizePolicy.setHeightForWidth(self.textBrowser.sizePolicy().hasHeightForWidth())
        self.textBrowser.setSizePolicy(sizePolicy)
        self.textBrowser.setObjectName("textBrowser")
        self.gridLayout.addWidget(self.textBrowser, 1, 0, 1, 1)
        self.listWidget_attachments = QtWidgets.QListWidget(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(1)
        sizePolicy.setHeightForWidth(self.listWidget_attachments.sizePolicy().hasHeightForWidth())
        self.listWidget_attachments.setSizePolicy(sizePolicy)
        self.listWidget_attachments.setMinimumSize(QtCore.QSize(0, 75))
        self.listWidget_attachments.setObjectName("listWidget_attachments")
        self.gridLayout.addWidget(self.listWidget_attachments, 2, 0, 1, 1)
        self.gridLayout_2.addLayout(self.gridLayout, 0, 1, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 691, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Почтослатель"))
        self.pushButton_ask_and_send.setText(_translate("MainWindow", "(2) Отправить письма всем выделенным людям"))
        self.pushButton_ask_and_send.setDisabled(True)
        self.pushButton_cancel_send.setText(_translate("MainWindow", "Отмена"))
        self.pushButton_cancel_send.setDisabled(True)
        self.pushButton_open_list_and_template.setText(_translate("MainWindow", "(1) Открыть *list.xlsx или ***text.html"))



###### batch_email_sender ######




def excepthook(excType, excValue, tracebackobj):
    traceback.print_tb(tracebackobj, excType, excValue)

sys.excepthook = excepthook


KEYRING_SERVICE = "batch_email_sender"
LAST_FROMMAIL = "pSxx7tJyvgz2tk"
LAST_MAILSERVER = "KjdsEYxeRaCk77"
LAST_FROMNAME = "FY8Btthta4n3ZF"
LAST_COPYLIST = "eqRyLeqKPatefP"
LAST_SAVEFLAG = "8AN43xqzGhZHUa"
LAST_PASSWORD = "uLkTjXd6BWa4tw"

EMAIL_REGEX = r"\s*([a-zA-Z0-9'_][a-zA-Z0-9'._+-]{,63}@[a-zA-Z0-9.-]{,254}[a-zA-Z0-9])\s*"



class Worker(QObject):
    sig_step = pyqtSignal(int, str)  # worker id, step description: emitted every step through work() loop
    sig_done = pyqtSignal(int)  # worker id: emitted at end of work()
    sig_mail_sent = pyqtSignal(int, int)
    sig_mail_error = pyqtSignal(int)

    def __init__(self, id: int, envelope):
        super().__init__()
        self.__id = id
        self.__abort = False
        self.envelope = envelope

    @pyqtSlot()
    def work(self):
        """
        Pretend this worker method does work that takes a long time. During this time, the thread's
        event loop is blocked, except if the application's processEvents() is called: this gives every
        thread (incl. main) a chance to process events, which in this sample means processing signals
        received from GUI (such as abort).
        """
        thread_name = QThread.currentThread().objectName()
        thread_id = int(QThread.currentThreadId())  # cast to int() is necessary
        self.sig_step.emit(self.__id, 'Running worker #{} from thread "{}" (#{})'.format(self.__id, thread_name, thread_id))

        while True:
            batch_sender_app.processEvents()  # this could cause change to self.__abort
            if self.__abort:
                self.sig_step.emit(self.__id, 'Worker #{} aborting work at step {}'.format(self.__id, step))
                break
            qt_mail_id, xls_mail_id = -1, -1
            try:
                mail = self.envelope.send_next()
                qt_mail_id, xls_mail_id, sent_to = mail['qt_id'], mail['xls_id'], mail['to_addrs']  # TODO здесь что-то грязно
            except StopIteration:
                break  # Это — победа
            except Exception as e:
                self.sig_step.emit(self.__id, f'Worker #{self.__id} error: {e!s}')
            if qt_mail_id >= 0:
                self.sig_step.emit(self.__id, f'Worker #{self.__id} sent to {sent_to}')
                self.sig_mail_sent.emit(qt_mail_id, xls_mail_id)
        self.sig_done.emit(self.__id)

    def abort(self):
        self.sig_step.emit(self.__id, 'Worker #{} notified to abort'.format(self.__id))
        self.__abort = True


class Extended_GUI(Ui_MainWindow, QObject):
    NUM_THREADS = 5
    USE_THREADS = None

    sig_abort_workers = pyqtSignal()

    def __init__(self, mainw):
        super().__init__()
        self.setupUi(mainw)
        self.CONFIG = ''
        self.template = ''
        self.xlsx_rows_list = ''
        self.parent = mainw
        self.pushButton_open_list_and_template.clicked.connect(self.open_xls_and_template)
        self.pushButton_ask_and_send.clicked.connect(self.send_msg)
        self.pushButton_cancel_send.clicked.connect(self.abort_workers)
        QThread.currentThread().setObjectName('main')  # threads can be named, useful for log output
        self.__workers_done = None
        self.__threads = None

    @pyqtSlot(int, str)
    def on_worker_step(self, worker_id: int, data: str):
        self.statusbar.showMessage('Worker #{}: {}'.format(worker_id, data))

    @pyqtSlot(int, int)
    def on_mail_sent(self, mail_widget_row_num: int, xls_row_number_ok: int):
        item = self.listWidget_emails.item(mail_widget_row_num)
        item.setBackground(QBrush(QColor("lightGreen")))  # Вах!
        item.setCheckState(False)
        try:
            set_ok(self.xls_name, xls_row_number_ok)
        except Exception as e:
            print(e)

    @pyqtSlot(int)
    def on_mail_error(self, mail_widget_row_num: int):
        item = self.listWidget_emails.item(mail_widget_row_num)
        item.setBackground(QBrush(QColor("lightRed")))  # Вах!

    @pyqtSlot(int)
    def on_worker_done(self, worker_id):
        self.statusbar.showMessage('worker #{} done'.format(worker_id))
        self.__workers_done += 1
        if self.__workers_done == self.USE_THREADS:
            self.statusbar.showMessage('No more workers active')
            self.pushButton_ask_and_send.setEnabled(True)
            self.pushButton_open_list_and_template.setEnabled(True)
            self.pushButton_cancel_send.setDisabled(True)
            QMessageBox.information(self.parent, 'OK', 'Все письма успешно отправлены!')

    @pyqtSlot()
    def abort_workers(self):
        self.sig_abort_workers.emit()
        self.statusbar.showMessage('Asking each worker to abort')
        for thread, worker in self.__threads:
            thread.quit()  # this will quit **as soon as thread event loop unblocks**
            thread.wait()  # <- so you need to wait for it to *actually* quit
        # even though threads have exited, there may still be messages on the main thread's
        # queue (messages that threads emitted before the abort):
        self.statusbar.showMessage('All threads exited')
        self.pushButton_ask_and_send.setEnabled(True)
        self.pushButton_open_list_and_template.setEnabled(True)
        self.pushButton_cancel_send.setDisabled(True)

    def show_email_attach(self, item):
        for i in range(self.listWidget_attachments.count()):
            if self.listWidget_attachments.item(i) == item and self.listWidget_attachments.item(i).text():
                if sys.platform.startswith('darwin'):
                    subprocess.call(('open', item.text()))
                elif os.name == 'nt':
                    os.startfile(item.text())
                elif os.name == 'posix':
                    subprocess.call(('xdg-open', item.text()))

    def read_list_and_template(self, filename):
        # filename — это либо имя excel'ника, либо имя html-шаблона. По имени определяем, что это, и определяем
        # оставшиеся имена
        self.template = None
        self.xlsx_rows_list = None
        self.pushButton_ask_and_send.setDisabled(True)
        filename = filename.lower()
        if filename.endswith('list.xlsx'):
            xls_name = filename
            template_name = filename.replace('list.xlsx', 'text.html')
        elif filename.endswith('text.html'):
            xls_name = filename.replace('text.html', 'list.xlsx')
            template_name = filename
        else:
            raise Exception(f'Нужно выбрать файл со списком ***list.xlsx или файл с шаблоном письма ***text.html')
        rows_list, bold_columns, template = rtv_table_and_template(xls_name, template_name)
        self.xls_name = xls_name
        self.template_name = template_name
        self.template = template
        self.xlsx_rows_list = rows_list
        if 'email' not in bold_columns:
            bold_columns.append('email')
        self.info_cols = bold_columns
        self.pushButton_ask_and_send.setEnabled(True)

    def update_preview_and_attaches_list(self, item):
        for i in range(self.listWidget_emails.count()):
            if self.listWidget_emails.item(i) == item:
                xlsx_row = self.xlsx_rows_list[i]
                self.textBrowser.setText(self.template.format(**xlsx_row))
                self.listWidget_attachments.clear()
                self.listWidget_attachments.addItems(xlsx_row['attach_list'])
                break

    def fill_widgets_with_emails(self):
        if not self.template or not self.xlsx_rows_list:
            return
        for xlsx_row in self.xlsx_rows_list:
            sample_row_text = ' '.join([xlsx_row[col] if col != 'email' else ', '.join(xlsx_row[col])
                                        for col in self.info_cols])
            item = QListWidgetItem(sample_row_text)
            item.xlsx_row = xlsx_row  # Именно отсюда мы возьмём данные для отправки
            item.setCheckState(Qt.Checked)
            if xlsx_row[OKOK] == OKOK:
                item.setCheckState(False)
                item.setBackground(QBrush(QColor("lightGreen")))  # Вах!
            self.listWidget_emails.addItem(item)
        self.listWidget_emails.itemClicked.connect(self.update_preview_and_attaches_list)
        self.listWidget_attachments.itemClicked.connect(self.show_email_attach)
        self.update_preview_and_attaches_list(self.listWidget_emails.item(0))

    def open_xls_and_template(self):
        filename, _ = QFileDialog.getOpenFileName(caption='Выберите список или шаблон', directory='',
                                                  options=QFileDialog.Options(),
                                                  filter="Список или шаблон (*list.xlsx *text.html);;Список (*list.xlsx);;Шаблон (*text.html);;All Files (*)")
        if not filename:
            return
        self.listWidget_emails.clear()
        self.listWidget_attachments.clear()
        self.textBrowser.clear()
        os.chdir(os.path.dirname(filename))
        try:
            self.read_list_and_template(filename)
        except Exception as e:
            QMessageBox.information(self.parent, 'OK', 'Ошибка: ' + str(e))
        self.fill_widgets_with_emails()

    def ask_login_and_create_connection(self):
        loginf = QDialog()
        diagui = Ui_Dialog()
        diagui.setupUi(loginf)

        last_frommail = keyring.get_password(KEYRING_SERVICE, LAST_FROMMAIL)
        last_fromname = keyring.get_password(KEYRING_SERVICE, LAST_FROMNAME)
        last_mailserver = keyring.get_password(KEYRING_SERVICE, LAST_MAILSERVER)
        last_copylist = keyring.get_password(KEYRING_SERVICE, LAST_COPYLIST)
        last_saveflag = keyring.get_password(KEYRING_SERVICE, LAST_SAVEFLAG)
        last_password = keyring.get_password(KEYRING_SERVICE, LAST_PASSWORD)
        diagui.line_email.setText(last_frommail or '')
        diagui.line_password.setText(last_password or '')
        diagui.line_sender.setText(last_fromname or '')
        diagui.line_smtpserver.setText(last_mailserver or 'smtp.googlemail.com')
        diagui.line_send_copy.setText(last_copylist or '')
        diagui.save_passw_cb.setCheckState(bool(last_saveflag))
        if loginf.exec_() == QDialog.Accepted:
            last_frommail = diagui.line_email.text()
            last_password = diagui.line_password.text()
            last_fromname = diagui.line_sender.text()
            last_mailserver = diagui.line_smtpserver.text()
            last_copylist = diagui.line_send_copy.text()
            last_saveflag = '1' if diagui.save_passw_cb.checkState() else ''
            keyring.set_password(KEYRING_SERVICE, LAST_FROMMAIL, last_frommail)
            keyring.set_password(KEYRING_SERVICE, LAST_FROMNAME, last_fromname)
            keyring.set_password(KEYRING_SERVICE, LAST_MAILSERVER, last_mailserver)
            keyring.set_password(KEYRING_SERVICE, LAST_COPYLIST, last_copylist)
            keyring.set_password(KEYRING_SERVICE, LAST_SAVEFLAG, last_saveflag)
            keyring.set_password(KEYRING_SERVICE, LAST_PASSWORD, last_password if last_saveflag else '')
            envelope = EmailEnvelope(smtp_server=last_mailserver,
                                             login=last_frommail,
                                             password=last_password,
                                             sender_addr=last_frommail,
                                             sender_name=last_fromname,
                                             copy_addrs=re.findall(EMAIL_REGEX, last_copylist))
            envelope.verify_credentials()
            return envelope
        else:
            raise ConnectionAbortedError('Необходимо ввести логин, пароль и т.д.')

    def start_threads_and_send_mails(self):
        mails_to_send = []
        for i in range(self.listWidget_emails.count()):
            item = self.listWidget_emails.item(i)
            if item.checkState():
                item.setSelected(True)
                xlsx_row = item.xlsx_row
                xlsx_row['QListWidgetIndex_WcCRve89'] = i  # Сохраняем номер строки, чтобы потом легко пометить её зелёным
                mails_to_send.append(item.xlsx_row)
        if not mails_to_send:
            msg = 'Ни одно письмо для отправки не выбрано'
            QMessageBox.information(self.parent, 'OK', msg)
            raise Exception(msg)
        # Создаём по отправляльщику на каждый worker
        # Равномерно распределяем на них на всех почту
        self.USE_THREADS = min(self.NUM_THREADS, len(mails_to_send))
        envelopes = [self.envelope.copy() for __ in range(self.USE_THREADS)]
        for i, xlsx_row in enumerate(mails_to_send):
            envelopes[i % self.USE_THREADS].add_mail_to_queue(recipients=xlsx_row['email'],
                                                              subject=xlsx_row['subject'],
                                                              html=self.template.format(**xlsx_row),
                                                              files=xlsx_row['attach_list'],
                                                              xls_id=xlsx_row[ORIGINAL_ROW_NUM],
                                                              qt_id=xlsx_row['QListWidgetIndex_WcCRve89'])
        # Лочим кнопки
        self.pushButton_ask_and_send.setDisabled(True)
        self.pushButton_open_list_and_template.setDisabled(True)
        self.pushButton_cancel_send.setEnabled(True)
        # Готовим worker'ов
        self.__workers_done = 0
        self.__threads = []
        for idx in range(self.USE_THREADS):
            worker = Worker(idx, envelopes[idx])
            thread = QThread()
            thread.setObjectName('thread_' + str(idx))
            self.__threads.append((thread, worker))  # need to store worker too otherwise will be gc'd
            worker.moveToThread(thread)

            # get progress messages from worker:
            worker.sig_step.connect(self.on_worker_step)
            worker.sig_done.connect(self.on_worker_done)
            worker.sig_mail_sent.connect(self.on_mail_sent)
            worker.sig_mail_error.connect(self.on_mail_error)

            # control worker:
            self.sig_abort_workers.connect(worker.abort)

            # get read to start worker:
            thread.started.connect(worker.work)
            thread.start()  # this will emit 'started' and start thread's event loop


    def send_msg(self):
        # Перед отправкой должен быть закружен шаблон и список
        if not self.xlsx_rows_list or not self.template:
            QMessageBox.warning(self.listWidget_emails.parent(), 'Ошибка', 'Сначала нужно открыть шаблон и список рассылки')
            raise Exception()

        self.envelope = None
        try:
            self.envelope = self.ask_login_and_create_connection()
        except ConnectionAbortedError:
            pass
        except Exception as e:
            msg = 'Не могу подключиться к серверу: ' + str(e)
            QMessageBox.warning(self.listWidget_emails.parent(), 'Ошибка', msg)
            raise Exception(msg)
        else:
            self.start_threads_and_send_mails()


if __name__ == '__main__':
    batch_sender_app = QApplication(sys.argv)
    main_window = QMainWindow()
    ui = Extended_GUI(main_window)
    main_window.show()
    sys.exit(batch_sender_app.exec_())
