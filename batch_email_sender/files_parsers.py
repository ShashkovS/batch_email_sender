import openpyxl
import os
import re

ONE_PLUS_ONE_FOR_HEADER = 2
OK_COLUMN = 1
OKOK = 'ok'
ORIGINAL_ROW_NUM = 'row_num_WcCRve89'
EMAIL_REGEX = r"\s*([a-zA-Z0-9'_][a-zA-Z0-9'._+-]{,63}@[a-zA-Z0-9.-]{,254}[a-zA-Z0-9])\s*"


def rtv_table(xls_name):
    try:
        xl_workbook = openpyxl.load_workbook(filename=xls_name, read_only=True, data_only=True)
    except FileNotFoundError:
        raise Exception('Файл ' + xls_name + ' не найден')
    xl_sheet_names = xl_workbook.sheetnames
    xl_sheet = xl_workbook[xl_sheet_names[0]]
    columns = []
    preview_columns = []
    row_iter = iter(xl_sheet.rows)
    for cell in next(row_iter):
        title = str(cell.value)
        if title:
            columns.append(title)  # cell.column
            if cell and cell.font and cell.font.bold:
                preview_columns.append(title)
    row_dict = {title: '' for title in columns}
    if 'email' not in row_dict:
        raise Exception('В файле {} обязательно должен быть столбец email'.format(xls_name))
    rows_list = []
    for rn, row in enumerate(row_iter, start=2):
        cur = row_dict.copy()
        cur[ORIGINAL_ROW_NUM] = rn
        for col_name, cell in zip(columns, row):
            cur[col_name] = str(cell.value).replace('None', '')
        cur['email'] = re.findall(EMAIL_REGEX, cur['email'])
        rows_list.append(cur)
    return rows_list, preview_columns


def set_ok(xls_name, row_num_real):
    while True:
        try:
            xl_workbook = openpyxl.load_workbook(filename=xls_name, read_only=False, data_only=False)
            break
        except FileNotFoundError:
            raise Exception('Файл ' + xls_name + ' не найден')
        except PermissionError:
            raise Exception(
                'Файл ' + xls_name + ' заблокирован. Сохраните и закойте его. В него будут вноситься отметки об успешности отправки')
            # continue
    xl_sheet_names = xl_workbook.get_sheet_names()
    xl_sheet = xl_workbook[xl_sheet_names[0]]
    xl_sheet.cell(row=row_num_real, column=OK_COLUMN).value = OKOK
    xl_workbook.save(filename=xls_name)


def rtv_template(template_name):
    try:
        with open(template_name, encoding='utf-8') as f:
            template = f.read()
    except FileNotFoundError:
        raise Exception('Файл с шаблоном ' + template_name + ' не найден')
    return template


def rtv_table_and_template(xls_name, template_name):
    rows_list, preview_columns = rtv_table(xls_name)
    template = rtv_template(template_name)
    if not rows_list:
        raise Exception('В файле ' + xls_name + ' не обнаружены строки с данными')
    first_data_row = rows_list[0]
    # Проверяем обязательные столбцы: ok, email, subject
    for col_name in ['ok', 'email', 'subject']:
        if col_name not in first_data_row:
            raise Exception('В таблице ' + xls_name + ' обязательно должен быть столбец ' + col_name)
    # Выкидываем строчки, в которых не заполнен email
    for i in range(len(rows_list) - 1, -1, -1):
        if not rows_list[i]['email']:
            rows_list.pop(i)
    # Проверяем, что есть всё, что указано в шаблоне
    try:
        template.format(**first_data_row)
    except KeyError as e:
        raise Exception('В таблице ' + xls_name + ' должен быть столбец ' + str(e)
                        + ', так как он упоминается в шаблоне ' + template_name)
    # Теперь проверяем существование всех вложений
    attach_cols = [key for key in first_data_row if key.startswith('attach')]
    for rn, row in enumerate(rows_list):
        row['attach_list'] = []
        for attach_key in attach_cols:
            attach_name = row[attach_key]
            if attach_name and not os.path.isfile(attach_name):
                raise Exception('В таблице ' + xls_name + ' в строчке ' +
                                str(rn + ONE_PLUS_ONE_FOR_HEADER) + ' в столбце ' + attach_key
                                + ' указано вложение "' + attach_name + '". Этот файл не найден')
            row['attach_list'].append(attach_name)
    # Проверили, что всё работает. Проверили, что вложения существуют
    return rows_list, preview_columns, template

# import os
# os.chdir(r'C:\Dropbox\repos\batch_email_sender')
# xls_name = 'send_list.xlsx'
# template_name = 'send_template.html'
# rows_list, preview_columns, template = rtv_table_and_template(xls_name, template_name)
# print(rows_list, preview_columns, template)

# result, preview_columns = rtv_table(xls_name)
# print(preview_columns)
# for row in result:
#     print(row)
